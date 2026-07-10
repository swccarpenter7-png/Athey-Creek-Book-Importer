from __future__ import annotations

import queue
import csv
import logging
import sys
import re
import shutil
import threading
import time
from copy import copy
from datetime import datetime
from pathlib import Path
from tkinter import BOTH, END, LEFT, RIGHT, W, X, BooleanVar, StringVar, Tk, filedialog, messagebox
from tkinter import ttk
from typing import Callable
from urllib.parse import urljoin, urlparse

import requests
from bs4 import BeautifulSoup, Tag
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

APP_NAME = "Athey Creek Book Importer"
APP_VERSION = "3.0.0"
BASE_URL = "https://atheycreek.com"
BOOKS_URL = f"{BASE_URL}/teachings/books"
USER_AGENT = f"Mozilla/5.0 (Windows NT 10.0; Win64; x64) {APP_NAME}/{APP_VERSION}"

BOOKS = [
    "Genesis", "Exodus", "Leviticus", "Numbers", "Deuteronomy", "Joshua", "Judges", "Ruth",
    "1 Samuel", "2 Samuel", "1 Kings", "2 Kings", "1 Chronicles", "2 Chronicles", "Ezra", "Nehemiah",
    "Esther", "Job", "Psalms", "Proverbs", "Ecclesiastes", "Song of Solomon", "Isaiah", "Jeremiah",
    "Lamentations", "Ezekiel", "Daniel", "Hosea", "Joel", "Amos", "Obadiah", "Jonah", "Micah",
    "Nahum", "Habakkuk", "Zephaniah", "Haggai", "Zechariah", "Malachi", "Matthew", "Mark", "Luke",
    "John", "Acts", "Romans", "1 Corinthians", "2 Corinthians", "Galatians", "Ephesians", "Philippians",
    "Colossians", "1 Thessalonians", "2 Thessalonians", "1 Timothy", "2 Timothy", "Titus", "Philemon",
    "Hebrews", "James", "1 Peter", "2 Peter", "1 John", "2 John", "3 John", "Jude", "Revelation",
]

SLUGS = {
    "1 Samuel": "1-samuel", "2 Samuel": "2-samuel", "1 Kings": "1-kings", "2 Kings": "2-kings",
    "1 Chronicles": "1-chronicles", "2 Chronicles": "2-chronicles",
    "Song of Solomon": "song-of-solomon", "1 Corinthians": "1-corinthians",
    "2 Corinthians": "2-corinthians", "1 Thessalonians": "1-thessalonians",
    "2 Thessalonians": "2-thessalonians", "1 Timothy": "1-timothy",
    "2 Timothy": "2-timothy", "1 Peter": "1-peter", "2 Peter": "2-peter",
    "1 John": "1-john", "2 John": "2-john", "3 John": "3-john",
}

HEADERS = [
    "Done", "Archive Source", "Testament", "Bible Order", "Book", "Lesson #",
    "Teaching Code", "Scripture", "Title", "Date", "Listen", "Listen URL", "Notes"
]


def application_folder() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def find_default_workbook() -> Path | None:
    candidates = []
    for path in application_folder().glob("*.xlsx"):
        if path.name.startswith("~$") or "_backup_" in path.stem.lower():
            continue
        candidates.append(path)
    if not candidates:
        return None
    non_templates = [p for p in candidates if "template" not in p.stem.lower()]
    return max(non_templates or candidates, key=lambda p: p.stat().st_mtime)


def write_import_report(book: str, workbook_path: Path, lessons: list[dict[str, str]], warnings: list[str], backup: Path) -> Path:
    reports = application_folder() / "reports"
    reports.mkdir(exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    report = reports / f"{book.replace(' ', '_')}_{stamp}.csv"
    with report.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle)
        writer.writerow(["Book", book])
        writer.writerow(["Workbook", str(workbook_path)])
        writer.writerow(["Backup", str(backup)])
        writer.writerow(["Imported Lessons", len(lessons)])
        writer.writerow(["Warnings", len(warnings)])
        writer.writerow([])
        writer.writerow(["Teaching Code", "Scripture", "Title", "Date", "Listen URL"])
        for lesson in lessons:
            writer.writerow([lesson["code"], lesson["scripture"], lesson["title"], lesson["date"], lesson["listen_url"]])
        if warnings:
            writer.writerow([])
            writer.writerow(["Warnings"] )
            for warning in warnings:
                writer.writerow([warning])
    return report


class ImporterError(RuntimeError):
    pass


def clean(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def make_session() -> requests.Session:
    session = requests.Session()
    session.headers.update({
        "User-Agent": USER_AGENT,
        "Accept-Language": "en-US,en;q=0.9",
    })
    return session


def fetch_soup(session: requests.Session, url: str, attempts: int = 3) -> BeautifulSoup:
    last_error: Exception | None = None
    for attempt in range(1, attempts + 1):
        try:
            response = session.get(url, timeout=35)
            response.raise_for_status()
            return BeautifulSoup(response.text, "html.parser")
        except requests.RequestException as exc:
            last_error = exc
            if attempt < attempts:
                time.sleep(attempt * 1.5)
    raise ImporterError(f"Could not open:\n{url}\n\n{last_error}")


def book_url(book: str) -> str:
    slug = SLUGS.get(book, book.lower().replace(" ", "-"))
    return f"{BASE_URL}/teachings/{slug}"


def lesson_urls(session: requests.Session, selected_book: str) -> list[str]:
    soup = fetch_soup(session, book_url(selected_book))
    found: list[str] = []
    seen: set[str] = set()
    pattern = re.compile(r"^/teachings/[A-Za-z]\d+-\d+$")

    for anchor in soup.find_all("a", href=True):
        absolute = urljoin(BASE_URL, anchor["href"])
        path = urlparse(absolute).path.rstrip("/")
        if pattern.fullmatch(path) and absolute not in seen:
            seen.add(absolute)
            found.append(absolute)

    if not found:
        raise ImporterError(
            f"No lesson links were found for {selected_book}.\n"
            "The Athey Creek website may have changed."
        )
    return found


def exact_heading(soup: BeautifulSoup) -> str:
    heading = soup.find("h2")
    if heading:
        value = clean(heading.get_text(" ", strip=True))
        if value:
            return value
    meta = soup.find("meta", property="og:title")
    if meta and meta.get("content"):
        return clean(meta["content"]).split("| Athey Creek Church")[0].strip()
    return ""


def exact_labeled_value(soup: BeautifulSoup, label: str) -> str:
    label_lower = label.lower()
    for tag in soup.find_all(["h4", "h5", "h6"]):
        if clean(tag.get_text(" ", strip=True)).lower() != label_lower:
            continue

        # The current site places the value after the label heading.
        for sibling in tag.next_siblings:
            if isinstance(sibling, Tag):
                value = clean(sibling.get_text(" ", strip=True))
                if value:
                    return value

        # Conservative fallback: stop at the next heading.
        current = tag.find_next()
        while current:
            if current is not tag and current.name in {"h4", "h5", "h6"}:
                break
            value = clean(current.get_text(" ", strip=True))
            if value and value.lower() != label_lower:
                return value
            current = current.find_next()
    return ""


def exact_date(soup: BeautifulSoup) -> str:
    # Current lesson pages show dates like "09 October 2013" in an h5.
    full_date = re.compile(
        r"^\d{1,2}\s+(January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{4}$",
        re.I,
    )
    short_date = re.compile(
        r"^(Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|Jun(?:e)?|Jul(?:y)?|"
        r"Aug(?:ust)?|Sep(?:tember)?|Oct(?:ober)?|Nov(?:ember)?|Dec(?:ember)?)\.?\s+\d{1,2},?\s+\d{4}$",
        re.I,
    )
    for tag in soup.find_all(["time", "h4", "h5", "h6", "p"]):
        raw = clean(tag.get("datetime") if tag.name == "time" else tag.get_text(" ", strip=True))
        if full_date.fullmatch(raw) or short_date.fullmatch(raw):
            return raw
    return ""


def audio_url(soup: BeautifulSoup) -> str:
    for selector in ("audio source[src]", "audio[src]", 'a[href*=".mp3"]'):
        tag = soup.select_one(selector)
        if tag:
            return urljoin(BASE_URL, tag.get("src") or tag.get("href"))
    for key in ("og:audio", "og:audio:url", "twitter:player:stream"):
        tag = soup.find("meta", property=key) or soup.find("meta", attrs={"name": key})
        if tag and tag.get("content"):
            return urljoin(BASE_URL, tag["content"])
    return ""


def parse_lesson(session: requests.Session, url: str) -> dict[str, str]:
    soup = fetch_soup(session, url)
    title = exact_heading(soup)
    scripture = exact_labeled_value(soup, "Book")
    date = exact_date(soup)
    code = urlparse(url).path.rstrip("/").split("/")[-1]
    audio = audio_url(soup)

    missing = [name for name, value in {
        "title": title, "scripture": scripture, "date": date
    }.items() if not value]
    if missing:
        raise ImporterError(
            f"Could not read {', '.join(missing)} from:\n{url}\n\n"
            "The site layout may have changed."
        )

    return {
        "code": code,
        "scripture": scripture,
        "title": title,
        "date": date,
        "listen_url": url,
        "audio_url": audio,
    }


def ensure_headers(ws) -> dict[str, int]:
    for col, header in enumerate(HEADERS, 1):
        ws.cell(1, col).value = header
    return {header: index for index, header in enumerate(HEADERS, 1)}


def copy_row_style(ws, source_row: int, target_row: int) -> None:
    for col in range(1, len(HEADERS) + 1):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        if source.has_style:
            target._style = copy(source._style)
        target.number_format = source.number_format
        target.alignment = copy(source.alignment)
        target.protection = copy(source.protection)
    if source_row in ws.row_dimensions:
        ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height


def old_user_data(ws, cols: dict[str, int], selected_book: str) -> dict[str, tuple[str, str]]:
    preserved: dict[str, tuple[str, str]] = {}
    for row in range(2, ws.max_row + 1):
        if clean(ws.cell(row, cols["Book"]).value).lower() != selected_book.lower():
            continue
        code = clean(ws.cell(row, cols["Teaching Code"]).value)
        if code:
            preserved[code] = (
                clean(ws.cell(row, cols["Done"]).value) or "☐",
                clean(ws.cell(row, cols["Notes"]).value),
            )
    return preserved


def update_dashboard_formulas(wb) -> None:
    if "Dashboard" not in wb.sheetnames:
        ws = wb.create_sheet("Dashboard")
    else:
        ws = wb["Dashboard"]

    ws["A1"] = "Athey Creek Through-the-Bible Progress"
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=16)
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    ws.merge_cells("A1:F1")
    ws["A1"].alignment = Alignment(horizontal="center")

    labels = [
        ("A3", "Metric"), ("B3", "Value"),
        ("A4", "Total Lessons"), ("A5", "Completed"),
        ("A6", "Remaining"), ("A7", "Overall Progress"),
        ("A8", "Old Testament Progress"), ("A9", "New Testament Progress"),
    ]
    for cell, value in labels:
        ws[cell] = value

    ws["B4"] = "=COUNTA('Master Index'!G:G)-1"
    ws["B5"] = '=COUNTIF(\'Master Index\'!A:A,"☑")'
    ws["B6"] = "=MAX(B4-B5,0)"
    ws["B7"] = "=IFERROR(B5/B4,0)"
    ws["B8"] = '=IFERROR(COUNTIFS(\'Master Index\'!C:C,"Old Testament",\'Master Index\'!A:A,"☑")/COUNTIF(\'Master Index\'!C:C,"Old Testament"),0)'
    ws["B9"] = '=IFERROR(COUNTIFS(\'Master Index\'!C:C,"New Testament",\'Master Index\'!A:A,"☑")/COUNTIF(\'Master Index\'!C:C,"New Testament"),0)'
    for cell in ("B7", "B8", "B9"):
        ws[cell].number_format = "0.0%"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18


def replace_book(
    workbook_path: Path,
    selected_book: str,
    lessons: list[dict[str, str]],
    preserve_progress: bool,
) -> Path:
    if workbook_path.suffix.lower() != ".xlsx":
        raise ImporterError("Please select an .xlsx workbook.")
    if not workbook_path.exists():
        raise ImporterError("The selected workbook does not exist.")

    backup = workbook_path.with_name(
        f"{workbook_path.stem}_backup_{datetime.now():%Y%m%d_%H%M%S}{workbook_path.suffix}"
    )
    shutil.copy2(workbook_path, backup)

    try:
        wb = load_workbook(workbook_path)
    except PermissionError as exc:
        raise ImporterError("Close the workbook in Excel, then run the import again.") from exc
    except Exception as exc:
        raise ImporterError(f"Could not open the workbook:\n{exc}") from exc

    if "Master Index" not in wb.sheetnames:
        raise ImporterError("The workbook must contain a sheet named 'Master Index'.")

    ws = wb["Master Index"]
    cols = ensure_headers(ws)
    preserved = old_user_data(ws, cols, selected_book) if preserve_progress else {}

    # Remove the old rows for only the selected book.
    for row in range(ws.max_row, 1, -1):
        if clean(ws.cell(row, cols["Book"]).value).lower() == selected_book.lower():
            ws.delete_rows(row)

    order = BOOKS.index(selected_book) + 1
    testament = "Old Testament" if order <= 39 else "New Testament"

    # Insert before the next Bible book already present.
    insert_at = ws.max_row + 1
    for row in range(2, ws.max_row + 1):
        try:
            existing_order = int(ws.cell(row, cols["Bible Order"]).value)
        except (TypeError, ValueError):
            continue
        if existing_order > order:
            insert_at = row
            break

    if lessons:
        ws.insert_rows(insert_at, amount=len(lessons))

    style_source_row = 2 if ws.max_row >= 2 else None
    for number, lesson in enumerate(lessons, 1):
        row = insert_at + number - 1
        if style_source_row and row != style_source_row:
            copy_row_style(ws, style_source_row, row)

        previous_done, previous_notes = preserved.get(lesson["code"], ("☐", ""))
        data = {
            "Done": previous_done,
            "Archive Source": "Current TTB Archive",
            "Testament": testament,
            "Bible Order": order,
            "Book": selected_book,
            "Lesson #": number,
            "Teaching Code": lesson["code"],
            "Scripture": lesson["scripture"],
            "Title": lesson["title"],
            "Date": lesson["date"],
            "Listen": "Listen",
            "Listen URL": lesson["listen_url"],
            "Notes": previous_notes,
        }

        for header, value in data.items():
            ws.cell(row, cols[header]).value = value

        listen_cell = ws.cell(row, cols["Listen"])
        listen_cell.hyperlink = lesson["listen_url"]
        listen_cell.style = "Hyperlink"

        url_cell = ws.cell(row, cols["Listen URL"])
        url_cell.hyperlink = lesson["listen_url"]
        url_cell.style = "Hyperlink"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:M{max(ws.max_row, 2)}"

    # Rebuild the table so filters and striped formatting cover all imported rows.
    for name in list(ws.tables.keys()):
        del ws.tables[name]
    if ws.max_row >= 2:
        table = Table(displayName="MasterIndexTable", ref=f"A1:M{ws.max_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)

    widths = {
        "A": 8, "B": 20, "C": 17, "D": 11, "E": 18, "F": 10, "G": 15,
        "H": 28, "I": 44, "J": 18, "K": 12, "L": 50, "M": 35,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows(min_row=2, max_col=len(HEADERS)):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    update_dashboard_formulas(wb)

    try:
        wb.save(workbook_path)
    except PermissionError as exc:
        raise ImporterError(
            "The workbook could not be saved. Close it in Excel and try again.\n\n"
            f"A safety backup was created at:\n{backup}"
        ) from exc

    return backup


def import_book(
    selected_book: str,
    workbook_path: Path,
    include_all: bool,
    preserve_progress: bool,
    progress: Callable[[str, int, int], None],
) -> tuple[int, Path]:
    session = make_session()
    urls = lesson_urls(session, selected_book)
    progress(f"Found {len(urls)} lesson pages. Reading official metadata…", 0, len(urls))

    lessons: list[dict[str, str]] = []
    warnings: list[str] = []
    seen_codes: set[str] = set()
    for index, url in enumerate(urls, 1):
        progress(f"Reading lesson {index} of {len(urls)}", index, len(urls))
        try:
            lesson = parse_lesson(session, url)
        except Exception as exc:
            warnings.append(f"{url}: {exc}")
            continue

        if lesson["code"] in seen_codes:
            warnings.append(f"Duplicate teaching code skipped: {lesson['code']}")
            continue
        seen_codes.add(lesson["code"])

        if include_all or lesson["title"].lower().startswith("through the bible"):
            lessons.append(lesson)

    if not lessons:
        raise ImporterError(
            f"No Through-the-Bible lessons were found for {selected_book}."
        )

    progress(f"Writing {len(lessons)} lessons to Excel…", len(urls), len(urls))
    backup = replace_book(workbook_path, selected_book, lessons, preserve_progress)
    report_path = write_import_report(selected_book, workbook_path, lessons, warnings, backup)
    return len(lessons), backup, report_path, warnings


class ImporterApp:
    def __init__(self) -> None:
        self.root = Tk()
        self.root.title(f"{APP_NAME} {APP_VERSION}")
        self.root.geometry("1080x720")
        self.root.minsize(940, 620)

        self.book = StringVar(value="Genesis")
        default_workbook = find_default_workbook()
        self.workbook = StringVar(value=str(default_workbook) if default_workbook else "")
        self.include_all = BooleanVar(value=False)
        self.preserve = BooleanVar(value=True)
        self.status = StringVar(value="Choose a Bible book and workbook.")
        self.messages: queue.Queue[tuple[str, object]] = queue.Queue()
        self.preview_lessons: list[dict[str, str]] = []

        self._build()
        self.root.after(100, self._poll_messages)

    def _build(self) -> None:
        frame = ttk.Frame(self.root, padding=18)
        frame.pack(fill=BOTH, expand=True)

        ttk.Label(frame, text=f"Athey Creek Book Importer v{APP_VERSION}", font=("Segoe UI", 18, "bold")).pack(anchor=W)
        ttk.Label(
            frame,
            text="Preview, verify, and import exact Scripture, title, date, code, and lesson links.",
        ).pack(anchor=W, pady=(2, 18))

        fields = ttk.Frame(frame)
        fields.pack(fill=X)

        ttk.Label(fields, text="Bible book:").grid(row=0, column=0, sticky=W, pady=6)
        book_box = ttk.Combobox(
            fields, textvariable=self.book, values=BOOKS, state="readonly", width=28
        )
        book_box.grid(row=0, column=1, sticky=W, padx=(10, 0), pady=6)

        ttk.Label(fields, text="Workbook:").grid(row=1, column=0, sticky=W, pady=6)
        ttk.Entry(fields, textvariable=self.workbook).grid(
            row=1, column=1, sticky="ew", padx=(10, 8), pady=6
        )
        ttk.Button(fields, text="Browse…", command=self._browse).grid(row=1, column=2, pady=6)
        fields.columnconfigure(1, weight=1)

        options = ttk.LabelFrame(frame, text="Options", padding=10)
        options.pack(fill=X, pady=14)
        ttk.Checkbutton(
            options,
            text="Preserve existing Done checkmarks and Notes for matching teaching codes",
            variable=self.preserve,
        ).pack(anchor=W)
        ttk.Checkbutton(
            options,
            text="Include all teachings on the book page (normally leave unchecked)",
            variable=self.include_all,
        ).pack(anchor=W, pady=(5, 0))

        actions = ttk.Frame(frame)
        actions.pack(fill=X, pady=(0, 12))
        self.preview_button = ttk.Button(actions, text="1. Preview Selected Book", command=self._start_preview)
        self.preview_button.pack(side=LEFT)
        self.import_button = ttk.Button(actions, text="2. Import Previewed Book", command=self._start_import, state="disabled")
        self.import_button.pack(side=LEFT, padx=8)
        self.verify_button = ttk.Button(actions, text="Verify Workbook Book", command=self._start_verify)
        self.verify_button.pack(side=LEFT)

        self.progress_bar = ttk.Progressbar(frame, mode="determinate")
        self.progress_bar.pack(fill=X)

        ttk.Label(frame, textvariable=self.status).pack(anchor=W, pady=(7, 8))

        preview_frame = ttk.LabelFrame(frame, text="Lesson Preview", padding=8)
        preview_frame.pack(fill=BOTH, expand=True)
        columns = ("lesson", "code", "scripture", "title", "date")
        self.preview_tree = ttk.Treeview(preview_frame, columns=columns, show="headings", height=12)
        for col, heading, width in [("lesson","Lesson",65),("code","Code",90),("scripture","Scripture",220),("title","Title",430),("date","Date",150)]:
            self.preview_tree.heading(col, text=heading)
            self.preview_tree.column(col, width=width, anchor=W)
        ybar = ttk.Scrollbar(preview_frame, orient="vertical", command=self.preview_tree.yview)
        xbar = ttk.Scrollbar(preview_frame, orient="horizontal", command=self.preview_tree.xview)
        self.preview_tree.configure(yscrollcommand=ybar.set, xscrollcommand=xbar.set)
        self.preview_tree.grid(row=0, column=0, sticky="nsew")
        ybar.grid(row=0, column=1, sticky="ns")
        xbar.grid(row=1, column=0, sticky="ew")
        preview_frame.rowconfigure(0, weight=1)
        preview_frame.columnconfigure(0, weight=1)

        log_frame = ttk.LabelFrame(frame, text="Activity", padding=8)
        log_frame.pack(fill=X, pady=(8,0))
        self.log = __import__("tkinter").Text(log_frame, height=7, wrap="word", state="disabled")
        self.log.pack(fill=X)

        ttk.Label(
            frame,
            text="Important: Close the workbook in Excel before importing. A timestamped backup is created automatically.",
        ).pack(anchor=W, pady=(10, 0))

    def _browse(self) -> None:
        chosen = filedialog.askopenfilename(
            title="Select Athey Creek workbook",
            filetypes=[("Excel workbooks", "*.xlsx")],
        )
        if chosen:
            self.workbook.set(chosen)

    def _append_log(self, line: str) -> None:
        self.log.configure(state="normal")
        self.log.insert(END, line + "\n")
        self.log.see(END)
        self.log.configure(state="disabled")

    def _set_busy(self, busy: bool) -> None:
        state = "disabled" if busy else "normal"
        self.preview_button.configure(state=state)
        self.verify_button.configure(state=state)
        self.import_button.configure(state="disabled" if busy or not self.preview_lessons else "normal")

    def _start_preview(self) -> None:
        self.preview_lessons = []
        for item in self.preview_tree.get_children():
            self.preview_tree.delete(item)
        self._set_busy(True)
        self.progress_bar["value"] = 0
        self.status.set("Reading official lesson pages…")
        threading.Thread(target=self._preview_worker, daemon=True).start()

    def _preview_worker(self) -> None:
        try:
            session = make_session()
            urls = lesson_urls(session, self.book.get())
            lessons, warnings, seen = [], [], set()
            for index, url in enumerate(urls, 1):
                self.messages.put(("progress", (f"Reading lesson {index} of {len(urls)}", index, len(urls))))
                try:
                    lesson = parse_lesson(session, url)
                except Exception as exc:
                    warnings.append(f"{url}: {exc}")
                    continue
                if lesson["code"] in seen:
                    warnings.append(f"Duplicate teaching code skipped: {lesson['code']}")
                    continue
                seen.add(lesson["code"])
                if self.include_all.get() or lesson["title"].lower().startswith("through the bible"):
                    lessons.append(lesson)
            if not lessons:
                raise ImporterError(f"No matching lessons were found for {self.book.get()}.")
            self.messages.put(("preview_done", (lessons, warnings)))
        except Exception as exc:
            self.messages.put(("error", str(exc)))

    def _start_import(self) -> None:
        path = Path(self.workbook.get().strip())
        if not path.exists():
            messagebox.showerror(APP_NAME, "Select a valid workbook first.")
            return
        if not self.preview_lessons:
            messagebox.showwarning(APP_NAME, "Preview the selected book before importing.")
            return
        self._set_busy(True)
        threading.Thread(target=self._import_previewed_worker, args=(path,), daemon=True).start()

    def _import_previewed_worker(self, path: Path) -> None:
        try:
            self.messages.put(("progress", ("Writing previewed lessons to Excel…", 1, 1)))
            backup = replace_book(path, self.book.get(), self.preview_lessons, self.preserve.get())
            report_path = write_import_report(self.book.get(), path, self.preview_lessons, [], backup)
            self.messages.put(("done", (self.book.get(), len(self.preview_lessons), path, backup, report_path, [])))
        except Exception as exc:
            self.messages.put(("error", str(exc)))

    def _start_verify(self) -> None:
        path = Path(self.workbook.get().strip())
        if not path.exists():
            messagebox.showerror(APP_NAME, "Select a valid workbook first.")
            return
        self._set_busy(True)
        threading.Thread(target=self._verify_worker, args=(path,), daemon=True).start()

    def _verify_worker(self, path: Path) -> None:
        try:
            session = make_session()
            urls = lesson_urls(session, self.book.get())
            live = {}
            for index, url in enumerate(urls, 1):
                self.messages.put(("progress", (f"Verifying lesson {index} of {len(urls)}", index, len(urls))))
                try:
                    lesson = parse_lesson(session, url)
                except Exception:
                    continue
                if self.include_all.get() or lesson["title"].lower().startswith("through the bible"):
                    live[lesson["code"]] = lesson
            wb = load_workbook(path, read_only=True, data_only=False)
            if "Master Index" not in wb.sheetnames:
                raise ImporterError("Workbook must contain a 'Master Index' tab.")
            ws = wb["Master Index"]
            cols = {clean(ws.cell(1,c).value): c for c in range(1, ws.max_column+1)}
            diffs=[]
            for r in range(2, ws.max_row+1):
                if clean(ws.cell(r, cols.get("Book",5)).value).lower()!=self.book.get().lower():
                    continue
                code=clean(ws.cell(r, cols.get("Teaching Code",7)).value)
                if not code or code not in live:
                    diffs.append(f"Row {r}: teaching code missing from live archive: {code}")
                    continue
                lesson=live[code]
                for field,header in [("scripture","Scripture"),("title","Title"),("date","Date")]:
                    current=clean(ws.cell(r, cols.get(header)).value)
                    if current!=clean(lesson[field]):
                        diffs.append(f"Row {r} {header}: workbook='{current}' | site='{lesson[field]}'")
            self.messages.put(("verify_done", diffs))
        except Exception as exc:
            self.messages.put(("error", str(exc)))

    def _run_worker(self, selected_book: str, path: Path, include_all: bool, preserve: bool) -> None:
        def report(text: str, current: int, total: int) -> None:
            self.messages.put(("progress", (text, current, total)))

        try:
            count, backup, report_path, warnings = import_book(
                selected_book, path, include_all, preserve, report
            )
            self.messages.put(("done", (selected_book, count, path, backup, report_path, warnings)))
        except Exception as exc:
            self.messages.put(("error", str(exc)))

    def _poll_messages(self) -> None:
        try:
            while True:
                event, payload = self.messages.get_nowait()
                if event == "progress":
                    text, current, total = payload
                    self.status.set(text)
                    self._append_log(text)
                    self.progress_bar["maximum"] = max(total, 1)
                    self.progress_bar["value"] = current
                elif event == "preview_done":
                    lessons, warnings = payload
                    self.preview_lessons = lessons
                    for number, lesson in enumerate(lessons, 1):
                        self.preview_tree.insert("", END, values=(number, lesson["code"], lesson["scripture"], lesson["title"], lesson["date"]))
                    self.status.set(f"Preview ready: {len(lessons)} lessons; {len(warnings)} warning(s).")
                    self._append_log(f"Preview ready for {self.book.get()}: {len(lessons)} lessons.")
                    self._set_busy(False)
                    if warnings:
                        self._append_log(f"Warnings: {len(warnings)}")
                elif event == "verify_done":
                    diffs = payload
                    self._set_busy(False)
                    if diffs:
                        self.status.set(f"Verification found {len(diffs)} difference(s).")
                        self._append_log("Verification differences:")
                        for line in diffs[:100]:
                            self._append_log(line)
                        messagebox.showwarning(APP_NAME, f"Verification found {len(diffs)} difference(s). See Activity for details.")
                    else:
                        self.status.set("Verification complete: no differences found.")
                        messagebox.showinfo(APP_NAME, "Verification complete. No differences were found for the selected book.")
                elif event == "done":
                    selected_book, count, path, backup, report_path, warnings = payload
                    self.status.set(f"Completed: {count} {selected_book} lessons imported.")
                    self._append_log(f"Completed. Workbook saved: {path}")
                    self._append_log(f"Backup created: {backup}")
                    self._append_log(f"Import report: {report_path}")
                    if warnings:
                        self._append_log(f"Warnings: {len(warnings)} (see report)")
                    self._set_busy(False)
                    messagebox.showinfo(
                        APP_NAME,
                        f"Completed!\n\n"
                        f"{count} {selected_book} lessons were imported.\n\n"
                        f"Workbook:\n{path}\n\n"
                        f"Backup:\n{backup}\n\n"
                        f"Import report:\n{report_path}\n\n"
                        f"Warnings: {len(warnings)}",
                    )
                elif event == "error":
                    self.status.set("Import failed.")
                    self._append_log(f"ERROR: {payload}")
                    self._set_busy(False)
                    messagebox.showerror(APP_NAME, str(payload))
        except queue.Empty:
            pass
        self.root.after(100, self._poll_messages)

    def run(self) -> None:
        self.root.mainloop()


if __name__ == "__main__":
    ImporterApp().run()
